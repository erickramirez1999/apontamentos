"""
Serviço de agrupamento e geração da planilha resumida.

Regras:
- Cada LINHA da saída = combinação única de (CLIENTE + EMPRESA classificada PISA/KING/TRIO)
- VENDEDOR e APELIDO = lista de todos separados por vírgula
- VALOR = soma dos títulos elegíveis dessa combinação
- ATRASO = maior atraso da combinação

Colunas de saída (modelo Pasta_de_Trabalho1.xlsx):
EMPRESA | PARCEIRO | NOME DO PARCEIRO | VENDEDOR | APELIDO | VALOR | ATRASO
"""
from __future__ import annotations

from io import BytesIO
from typing import Optional

import pandas as pd

from src.servicos.filtragem_sankhya import classificar_empresa


def agrupar_para_planilha_resumo(df_validos: pd.DataFrame) -> pd.DataFrame:
    """
    Agrupa títulos válidos em: 1 linha por (Cliente + Empresa).

    Retorna DataFrame com colunas:
    EMPRESA | PARCEIRO | NOME DO PARCEIRO | VENDEDOR | APELIDO | VALOR | ATRASO
    """
    if df_validos.empty:
        return pd.DataFrame(columns=[
            "EMPRESA", "PARCEIRO", "NOME DO PARCEIRO",
            "VENDEDOR", "APELIDO", "VALOR ", "ATRASO "
        ])

    # Adicionar coluna de empresa classificada
    df = df_validos.copy()
    df["empresa_lle"] = df.apply(
        lambda r: classificar_empresa(r["Empresa"], r["Vendedor"]),
        axis=1
    )

    # Agrupar
    linhas = []
    for (cod_parceiro, empresa), sub in df.groupby(["Parceiro", "empresa_lle"]):
        nome = sub["Nome Parceiro (Parceiro)"].iloc[0]

        # Vendedores únicos (preservando ordem de aparição)
        vendedores_unicos = []
        vendedores_set = set()
        for v in sub["Vendedor"]:
            if pd.notna(v):
                vi = int(v)
                if vi not in vendedores_set:
                    vendedores_unicos.append(vi)
                    vendedores_set.add(vi)

        apelidos_unicos = []
        apelidos_set = set()
        for a in sub["Apelido"]:
            if pd.notna(a):
                ai = str(a).strip()
                if ai and ai not in apelidos_set:
                    apelidos_unicos.append(ai)
                    apelidos_set.add(ai)

        valor_total = float(sub["Vlr do Desdobramento"].sum())
        atraso_max = int(sub["Atraso (dias)"].max())

        linhas.append({
            "EMPRESA": empresa,
            "PARCEIRO": int(cod_parceiro),
            "NOME DO PARCEIRO": nome,
            "VENDEDOR": ", ".join(str(v) for v in vendedores_unicos),
            "APELIDO": ", ".join(apelidos_unicos),
            "VALOR ": round(valor_total, 2),
            "ATRASO ": atraso_max,
        })

    df_out = pd.DataFrame(linhas)
    # Ordenar por VALOR desc (maiores primeiro)
    df_out = df_out.sort_values("VALOR ", ascending=False).reset_index(drop=True)
    return df_out


def gerar_excel_resumo(df_resumo: pd.DataFrame) -> bytes:
    """
    Gera Excel (.xlsx) a partir do DataFrame resumido.
    Retorna bytes prontos pra download.

    Pinta as linhas por empresa pra confirmação visual:
      - PISA → azul claro
      - KING → amarelo claro
      - TRIO → verde claro
    """
    from openpyxl.styles import PatternFill, Font

    # Cores por empresa (mesmas usadas na prévia da tela)
    CORES_EMPRESA = {
        "PISA": "D6E4FF",  # azul claro
        "KING": "FFF5CC",  # amarelo claro
        "TRIO": "D6F5D6",  # verde claro
    }

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_resumo.to_excel(writer, sheet_name="Planilha1", index=False)

        ws = writer.sheets["Planilha1"]

        # Ajustar largura das colunas
        larguras = {
            "A": 10,   # EMPRESA
            "B": 12,   # PARCEIRO
            "C": 50,   # NOME
            "D": 25,   # VENDEDOR
            "E": 35,   # APELIDO
            "F": 14,   # VALOR
            "G": 10,   # ATRASO
        }
        for col, larg in larguras.items():
            ws.column_dimensions[col].width = larg

        # Pintar cabeçalho (linha 1) em cinza com negrito
        header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Achar índice da coluna EMPRESA
        col_empresa_idx = None
        for idx, cell in enumerate(ws[1], start=1):
            if str(cell.value).strip().upper() == "EMPRESA":
                col_empresa_idx = idx
                break

        # Pintar cada linha de dados conforme a empresa
        if col_empresa_idx is not None:
            for row_idx in range(2, ws.max_row + 1):
                empresa_cell = ws.cell(row=row_idx, column=col_empresa_idx)
                empresa = str(empresa_cell.value).strip().upper()
                cor = CORES_EMPRESA.get(empresa)
                if cor:
                    fill = PatternFill(start_color=cor, end_color=cor, fill_type="solid")
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=row_idx, column=col).fill = fill

    return output.getvalue()
